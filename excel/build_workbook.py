import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice as DrawColorChoice
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
GOLD = "B08D57"
LIGHT = "EAF1F8"
WHITE = "FFFFFF"

df = pd.read_csv("../data/freight_shipments_clean.csv")

wb = Workbook()

# ---------------------------------------------------------------
# Sheet: Data
# ---------------------------------------------------------------
ws_data = wb.active
ws_data.title = "Data"
cols = ["ShipmentID", "ShipDate", "Route", "Carrier", "WeightKg", "Cost",
        "TransitDays", "OnTime", "CostPerKg", "Month", "OnTimeFlag"]
ws_data.append(cols)
for c in ws_data[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for _, row in df.iterrows():
    ws_data.append([row[c] for c in cols])
for i in range(1, len(cols) + 1):
    ws_data.column_dimensions[get_column_letter(i)].width = 13
n_rows = len(df) + 1

# ---------------------------------------------------------------
# Sheet: Route_Carrier_Summary (formula-driven pivot)
# ---------------------------------------------------------------
ws_s = wb.create_sheet("Route_Carrier_Summary")
combos = df.groupby(["Route", "Carrier"]).size().reset_index()[["Route", "Carrier"]].values.tolist()
ws_s.append(["Route", "Carrier", "Shipments", "AvgCostPerKg", "OnTimePct", "TotalCost"])
for c in ws_s[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, (route, carrier) in enumerate(combos):
    r = i + 2
    ws_s.cell(row=r, column=1, value=route)
    ws_s.cell(row=r, column=2, value=carrier)
    ws_s.cell(row=r, column=3, value=f'=COUNTIFS(Data!$C$2:$C${n_rows},A{r},Data!$D$2:$D${n_rows},B{r})')
    ws_s.cell(row=r, column=4, value=f'=ROUND(AVERAGEIFS(Data!$I$2:$I${n_rows},Data!$C$2:$C${n_rows},A{r},Data!$D$2:$D${n_rows},B{r}),3)')
    ws_s.cell(row=r, column=5, value=f'=ROUND(AVERAGEIFS(Data!$K$2:$K${n_rows},Data!$C$2:$C${n_rows},A{r},Data!$D$2:$D${n_rows},B{r})*100,1)')
    ws_s.cell(row=r, column=6, value=f'=ROUND(SUMIFS(Data!$F$2:$F${n_rows},Data!$C$2:$C${n_rows},A{r},Data!$D$2:$D${n_rows},B{r}),0)')
last_s_row = len(combos) + 1
for i in range(1, 7):
    ws_s.column_dimensions[get_column_letter(i)].width = 22

# Carrier-level efficiency frontier table (for the scatter chart)
ws_s.cell(row=1, column=8, value="Carrier").font = Font(bold=True, color=WHITE, name="Arial")
ws_s.cell(row=1, column=9, value="AvgCostPerKg").font = Font(bold=True, color=WHITE, name="Arial")
ws_s.cell(row=1, column=10, value="OnTimePct").font = Font(bold=True, color=WHITE, name="Arial")
for c in (8, 9, 10):
    ws_s.cell(row=1, column=c).fill = PatternFill("solid", fgColor=GOLD)
carriers = sorted(df["Carrier"].unique())
for i, carrier in enumerate(carriers):
    r = i + 2
    ws_s.cell(row=r, column=8, value=carrier)
    ws_s.cell(row=r, column=9, value=f'=ROUND(AVERAGEIF(Data!$D$2:$D${n_rows},H{r},Data!$I$2:$I${n_rows}),3)')
    ws_s.cell(row=r, column=10, value=f'=ROUND(AVERAGEIF(Data!$D$2:$D${n_rows},H{r},Data!$K$2:$K${n_rows})*100,1)')
last_carrier_row = len(carriers) + 1
for i in (8, 9, 10):
    ws_s.column_dimensions[get_column_letter(i)].width = 20

# ---------------------------------------------------------------
# Sheet: Savings_Scenario
# ---------------------------------------------------------------
ws_sc = wb.create_sheet("Savings_Scenario")
ws_sc["A1"] = "Route Reallocation Scenario: EU Direct (Dublin-Cherbourg)"
ws_sc["A1"].font = Font(bold=True, size=13, color=NAVY, name="Arial")
labels = ["Carrier", "Total Weight (kg)", "Actual Cost (€)", "Cost/kg if moved to Celtic Sealink rate",
          "Projected Cost at Celtic Rate (€)", "Projected Saving (€)", "Projected Saving (%)"]
for i, lbl in enumerate(labels):
    ws_sc.cell(row=3 + i, column=1, value=lbl).font = Font(bold=True, name="Arial")

ws_sc["B3"] = "Atlantic Cargo Ferries"
ws_sc["B4"] = '=SUMIFS(Data!$E$2:$E$' + str(n_rows) + ',Data!$C$2:$C$' + str(n_rows) + ',"EU Direct (Dublin-Cherbourg)",Data!$D$2:$D$' + str(n_rows) + ',"Atlantic Cargo Ferries")'
ws_sc["B5"] = '=SUMIFS(Data!$F$2:$F$' + str(n_rows) + ',Data!$C$2:$C$' + str(n_rows) + ',"EU Direct (Dublin-Cherbourg)",Data!$D$2:$D$' + str(n_rows) + ',"Atlantic Cargo Ferries")'
ws_sc["B6"] = '=SUMIFS(Data!$F$2:$F$' + str(n_rows) + ',Data!$C$2:$C$' + str(n_rows) + ',"EU Direct (Dublin-Cherbourg)",Data!$D$2:$D$' + str(n_rows) + ',"Celtic Sealink")/SUMIFS(Data!$E$2:$E$' + str(n_rows) + ',Data!$C$2:$C$' + str(n_rows) + ',"EU Direct (Dublin-Cherbourg)",Data!$D$2:$D$' + str(n_rows) + ',"Celtic Sealink")'
ws_sc["B7"] = "=ROUND(B4*B6,0)"
ws_sc["B8"] = "=ROUND(B5-B7,0)"
ws_sc["B9"] = "=ROUND(B8/B5*100,1)"
for r in range(3, 10):
    ws_sc.cell(row=r, column=2).font = Font(size=12, color=GOLD, bold=True, name="Arial")
ws_sc.column_dimensions["A"].width = 40
ws_sc.column_dimensions["B"].width = 24

# ---------------------------------------------------------------
# Sheet: Dashboard
# ---------------------------------------------------------------
ws_d = wb.create_sheet("Dashboard", 0)
ws_d.sheet_view.showGridLines = False

ws_d.merge_cells("B2:K2")
ws_d["B2"] = "DUBLIN PORT FREIGHT & LOGISTICS COST DASHBOARD"
ws_d["B2"].font = Font(bold=True, size=19, color=NAVY, name="Arial")
ws_d.merge_cells("B3:K3")
ws_d["B3"] = "3 Routes · 3 Carriers · 2024-2025"
ws_d["B3"].font = Font(italic=True, size=12, color=GOLD, name="Arial")

def kpi_card(ws, col, label, formula, fmt="#,##0"):
    col_letter = get_column_letter(col)
    ws.merge_cells(f"{col_letter}5:{get_column_letter(col+1)}5")
    ws[f"{col_letter}5"] = label
    ws[f"{col_letter}5"].font = Font(bold=True, color=WHITE, size=11, name="Arial")
    ws[f"{col_letter}5"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"{col_letter}5"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"{col_letter}6:{get_column_letter(col+1)}7")
    cell = ws[f"{col_letter}6"]
    cell.value = formula
    cell.font = Font(bold=True, size=19, color=GOLD, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    for rr in (5, 6, 7):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

kpi_card(ws_d, 2, "TOTAL FREIGHT SPEND", f"=ROUND(SUM(Data!F2:F{n_rows}),0)", '#,##0" €"')
kpi_card(ws_d, 4, "TOTAL SHIPMENTS", f"=COUNTA(Data!A2:A{n_rows})", "#,##0")
kpi_card(ws_d, 6, "AVG ON-TIME %", f"=ROUND(AVERAGE(Data!K2:K{n_rows})*100,1)", '0.0"%"')
kpi_card(ws_d, 8, "PROJECTED SAVING", "=Savings_Scenario!B8", '#,##0" €"')
kpi_card(ws_d, 10, "SAVING %", "=Savings_Scenario!B9", '0.0"%"')

ws_d.row_dimensions[6].height = 22
ws_d.row_dimensions[7].height = 22

# Cost/kg by route+carrier bar chart
bar = BarChart()
bar.title = "Avg Cost/kg by Route & Carrier"
bar.style = 10
bar.y_axis.title = "€ per kg"
data = Reference(ws_s, min_col=4, min_row=1, max_row=last_s_row)
cats = Reference(ws_s, min_col=2, min_row=2, max_row=last_s_row)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 16, 9
ws_d.add_chart(bar, "B10")

# Efficiency frontier scatter: cost/kg (x) vs on-time% (y) per carrier
scatter = ScatterChart()
scatter.title = "Efficiency Frontier: Cost vs. Reliability by Carrier"
scatter.x_axis.title = "Avg Cost per kg (€)"
scatter.y_axis.title = "On-Time %"
xvalues = Reference(ws_s, min_col=9, min_row=2, max_row=last_carrier_row)
yvalues = Reference(ws_s, min_col=10, min_row=2, max_row=last_carrier_row)
series = Series(yvalues, xvalues, title="Carriers")
series.marker = Marker(symbol="circle", size=12)
series.marker.graphicalProperties.solidFill = "1F3864"
series.marker.graphicalProperties.line.solidFill = "1F3864"
series.graphicalProperties.line.noFill = True
scatter.series.append(series)
scatter.width, scatter.height = 16, 9
ws_d.add_chart(scatter, "B29")

for i in range(1, 12):
    ws_d.column_dimensions[get_column_letter(i)].width = 16
ws_d.page_setup.orientation = "landscape"
ws_d.page_setup.fitToWidth = 1
ws_d.page_setup.fitToHeight = 0
ws_d.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Dublin_Freight_Cost_Dashboard.xlsx")
print("saved")
